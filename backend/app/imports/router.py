from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
import csv
import io
from typing import List, Dict, Any

from openpyxl import load_workbook

from app.database import get_db
from app.utils.dependencies import get_current_user
from app.models.usuario import Usuario
from app.services import libro_service
from app.predictions.service import normalize_header, normalize_inventory_row, predict_inventory_from_rows

router = APIRouter()

PRODUCT_FIELD_ALIASES = {
    "titulo": ("titulo", "producto", "product", "nombre", "name", "descripcion_corta"),
    "autor": ("autor", "marca", "brand", "proveedor", "supplier", "fabricante"),
    "isbn": ("isbn", "sku", "codigo", "codigo_producto", "code"),
    "categoria": ("categoria", "category", "rubro", "linea", "familia"),
    "stock_total": ("stock_total", "stock", "inventory", "inventario", "cantidad", "quantity", "existencias"),
    "descripcion": ("descripcion", "description", "detalle", "observaciones"),
}


def _normalize_header(value: Any) -> str:
    return normalize_header(value)


def _extract_rows_from_csv(content: bytes) -> List[Dict[str, Any]]:
    text = content.decode('utf-8', errors='ignore')
    reader = csv.DictReader(io.StringIO(text))
    return [{_normalize_header(key): value for key, value in row.items()} for row in reader]


def _extract_rows_from_excel(content: bytes) -> List[Dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    rows = []
    headers = []

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_index == 1:
            headers = [_normalize_header(cell) for cell in row]
            continue

        if not any(cell not in (None, "") for cell in row):
            continue

        data = {}
        for index, value in enumerate(row):
            if index >= len(headers):
                break
            data[headers[index]] = value

        rows.append(data)

    workbook.close()
    return rows


def _first_alias(row: Dict[str, Any], aliases: tuple[str, ...], default: Any = None) -> Any:
    normalized = {_normalize_header(key): value for key, value in row.items()}
    for alias in aliases:
        value = normalized.get(alias)
        if value not in (None, ""):
            return value
    return default


def _to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        if isinstance(value, str):
            value = value.strip().replace(",", ".")
        return float(value)
    except (TypeError, ValueError):
        return None


def _profile_dataset(rows: List[Dict[str, Any]], normalized_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    columns = list(rows[0].keys()) if rows else []
    numeric = {}
    for column in columns:
        values = [_to_float(row.get(column)) for row in rows]
        numbers = [value for value in values if value is not None]
        if numbers:
            avg = sum(numbers) / len(numbers)
            recent = numbers[-3:] if len(numbers) >= 3 else numbers
            predicted = sum(recent) / len(recent)
            previous = sum(numbers[:-1]) / len(numbers[:-1]) if len(numbers) > 1 else numbers[-1]
            delta = numbers[-1] - previous
            if abs(delta) <= max(1, abs(previous) * 0.03):
                trend = "estable"
            else:
                trend = "sube" if delta > 0 else "baja"
            numeric[column] = {
                "count": len(numbers),
                "sum": round(sum(numbers), 2),
                "avg": round(avg, 2),
                "min": round(min(numbers), 2),
                "max": round(max(numbers), 2),
                "values": numbers[:100],
            }
            numeric.setdefault("_predictions", {})[column] = {
                "predicted": round(predicted, 2),
                "trend": trend,
                "classification": "bajo" if predicted <= avg * 0.85 else "alto" if predicted >= avg * 1.15 else "normal",
            }

    products = {row.get("product") for row in normalized_rows if row.get("product")}
    total_stock = sum(float(row.get("stock") or 0) for row in normalized_rows)
    total_outgoing = sum(float(row.get("quantity") or 0) for row in normalized_rows)
    total_incoming = sum(float(row.get("incoming") or 0) for row in normalized_rows)
    low_stock = sum(
        1
        for row in normalized_rows
        if float(row.get("stock") or 0) <= max(float(row.get("minimum_stock") or 0), 50)
    )

    return {
        "columns": columns,
        "numeric": {key: value for key, value in numeric.items() if key != "_predictions"},
        "column_predictions": numeric.get("_predictions", {}),
        "totals": {
            "rows": len(rows),
            "products": len(products) or len(rows),
            "stock": int(round(total_stock)),
            "outgoing": int(round(total_outgoing)),
            "incoming": int(round(total_incoming)),
            "low_stock": low_stock,
        },
    }


@router.post("/products/csv")
def import_products_csv(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV supported for now")

    content = file.file.read().decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    created = 0
    errors = []
    for i, row in enumerate(reader, start=1):
        try:
            data = {
                'titulo': _first_alias(row, PRODUCT_FIELD_ALIASES["titulo"]),
                'autor': _first_alias(row, PRODUCT_FIELD_ALIASES["autor"], "Sin proveedor"),
                'isbn': _first_alias(row, PRODUCT_FIELD_ALIASES["isbn"]),
                'categoria': _first_alias(row, PRODUCT_FIELD_ALIASES["categoria"], 'Uncategorized'),
                'stock_total': int(float(_first_alias(row, PRODUCT_FIELD_ALIASES["stock_total"], 1))),
                'descripcion': _first_alias(row, PRODUCT_FIELD_ALIASES["descripcion"]),
            }
            libro_service.crear_libro(db, type('R',(),{'model_dump':lambda self, **kw: data})(), client_id=current_user.client_id)
            created += 1
        except Exception as e:
            errors.append({'line': i, 'error': str(e), 'row': row})

    return {'created': created, 'errors': errors}


@router.post("/inventory/upload-and-predict")
def upload_inventory_and_predict(file: UploadFile = File(...), current_user: Usuario = Depends(get_current_user)):
    if not file.filename.lower().endswith(('.csv', '.xlsx')):
        raise HTTPException(status_code=400, detail="Only CSV or Excel files are supported")

    content = file.file.read()
    if file.filename.lower().endswith('.xlsx'):
        rows = _extract_rows_from_excel(content)
    else:
        rows = _extract_rows_from_csv(content)

    if not rows:
        raise HTTPException(status_code=400, detail="No rows found in file")

    prediction = predict_inventory_from_rows(rows)
    normalized_rows = [normalize_inventory_row(row) for row in rows]
    profile = _profile_dataset(rows, normalized_rows)
    return {
        'imported_rows': len(rows),
        'prediction': prediction,
        'classification': prediction['classification'],
        'columns': profile["columns"],
        'profile': profile,
        'rows': normalized_rows,
        'raw_rows': rows,
    }
